"""
sheet.py
Functions related to reading and writing data in the Excel file
"""
# Import dependencies
from openpyxl import load_workbook
from openpyxl.workbook import workbook

current_workbook = 1

"""
Prepare Book
Prepare workbook, create default sheet, apply settings
Args:-
    :mode - (int)
    :project_name - (str)
    :project_data_folder - (str)
    
    mode 0 - create new
    mode 1 - load existing
"""


def prepare_book(mode: int, project_data_folder: str, project_name: str, workbook_path: str = None):
    try:
        if mode == 0:
            wb = workbook.Workbook()
            wb.create_sheet(f"{project_name}{current_workbook}")
            wb.save(project_data_folder)
            print("Created")
            return wb

        elif mode == 1:
            wb = load_workbook(workbook_path)
            print("loaded")
            return wb

    except Exception as workbook_error:
        print(workbook_error)


"""
Add rows
Setup workbook by adding rows to it

Args:
    rows (list) 
"""


def add_rows(rows: list, wb: workbook.Worksheet):
    try:
        rows_len = len(rows)
        wb.insert_rows(rows_len)

    except Exception as error:
        print(error)
